#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Automatizador de Extratos de Cartão de Crédito
Versão com conciliação automática e manual de fornecedores via API Omie.
Suporte para PDF do Santander (texto e imagem)

MAPEAMENTO DE COLUNAS:
- Coluna C: Fornecedor (descrição da compra)
- Coluna D: Categoria (sempre "Cartão de Credito")
- Coluna E: Conta Corrente (informada pelo usuário)
- Coluna F: Valor da Conta (valor da transação)
- Coluna J: Data de Registro (data da compra)
- Coluna K: Data de Vencimento (informada pelo usuário)

Inserção inicia na linha 6.
"""

import os
import re
import sys
import shutil
import json
import html
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from typing import List, Dict, Tuple, Optional
import xml.etree.ElementTree as ET

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    import PyPDF2
    import pdfplumber
    import requests
    from fuzzywuzzy import fuzz
    from omie_api import get_clientes_as_fornecedores, get_categorias
    import pytesseract
    from PIL import Image
    import fitz  # PyMuPDF
except ImportError as e:
    messagebox.showerror(
        "Erro",
        f"Erro: Biblioteca necessária não encontrada: {e}\n\n"
        "Instale as dependências com:\n"
        "pip install pandas openpyxl PyPDF2 pdfplumber requests fuzzywuzzy pytesseract pillow PyMuPDF\n\n"
        "Também instale o Tesseract OCR: https://github.com/tesseract-ocr/tesseract"
    )
    sys.exit(1)

class ExtractProcessor:
    """
    Classe principal para processar extratos de cartão de crédito
    e conciliar com a Omie.
    """
    def __init__(self):
        self.supported_banks = ["Sicoob", "Banco do Brasil", "Caixa", "Itaú", "Santander"]
        self.file_formats = {
            "Sicoob": "OFX",
            "Banco do Brasil": "PDF",
            "Caixa": "PDF",
            "Itaú": "PDF",
            "Santander": "PDF"  # Alterado de OFX para PDF
        }
        self.omie_suppliers = []
        self.omie_categories = []

    def _load_credentials(self, client_name: str) -> Optional[Dict]:
        file_path = f"credenciais/{client_name.replace(' ', '_').lower()}.json"
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            messagebox.showerror("Erro de Credenciais", f"Arquivo de credenciais para '{client_name}' não encontrado.")
            return None
        except json.JSONDecodeError:
            messagebox.showerror("Erro de Credenciais", "Arquivo de credenciais inválido.")
            return None
    
    def _process_and_reconcile(self, bank: str, extract_file: str, client: str) -> Optional[List[Dict]]:
        transactions = self._process_extract(extract_file, self.file_formats[bank], bank)
        
        if not transactions:
            messagebox.showinfo("Aviso", "Nenhuma transação encontrada no extrato.")
            return None

        credentials = self._load_credentials(client)
        if not credentials:
            return None
        
        app_key = credentials.get("app_key")
        app_secret = credentials.get("app_secret")
        
        if not all([app_key, app_secret]):
            messagebox.showerror("Erro", "Credenciais de API incompletas.")
            return None

        self.omie_suppliers = get_clientes_as_fornecedores(app_key, app_secret)
        self.omie_categories = get_categorias(app_key, app_secret)

        if not self.omie_suppliers:
            messagebox.showinfo("Aviso", "Nenhum fornecedor encontrado na Omie para este cliente.")
            return transactions
        
        for transaction in transactions:
            description = transaction['fornecedor']
            best_match = None
            highest_score = 0
            
            for supplier in self.omie_suppliers:
                omie_name = supplier.get('nome_fantasia') or supplier.get('razao_social')
                if omie_name:
                    score = fuzz.ratio(description.lower(), omie_name.lower())
                    if score > highest_score:
                        highest_score = score
                        best_match = omie_name

            if highest_score > 80:
                transaction['fornecedor_omie'] = best_match
            else:
                transaction['fornecedor_omie'] = ""
                
        return transactions

    def process_and_save(self, transactions: List[Dict], account: str, due_date: str) -> str:
        """
        Processa e salva os dados na planilha final.
        """
        # NOVO CAMINHO: A linha abaixo foi alterada para o novo local da planilha
        base_file = "C:\\Bitrix24\\Aurora Hotel\\Automação\\Omie_Contas_Pagar_v1_1_5.xlsx"
        
        if not os.path.exists(base_file):
            return f"ERRO: Arquivo base '{base_file}' não encontrado!"

        new_file_path = self._create_new_excel_file(base_file)
        if new_file_path:
            self._insert_into_excel(new_file_path, transactions, account, due_date)
            return f"✅ Processamento concluído! {len(transactions)} transações inseridas.\n\nArquivo atualizado: {new_file_path}"
        else:
            return "Erro ao criar a nova planilha."

    def _create_new_excel_file(self, base_file: str) -> str:
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_file_name = f"Omie_Contas_Pagar_Atualizada_{timestamp}.xlsx"
            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
            new_file_path = os.path.join(desktop_path, new_file_name)
            shutil.copy2(base_file, new_file_path)
            return new_file_path
        except Exception as e:
            print(f"Erro ao criar nova planilha: {e}")
            return ""

    def _process_extract(self, file_path: str, file_format: str, bank: str) -> List[Dict]:
        if file_format == "OFX":
            return self._process_ofx(file_path)
        elif file_format == "PDF":
            return self._process_pdf(file_path, bank)
        elif file_format == "Excel":
            return self._process_excel(file_path, bank)
        else:
            messagebox.showerror("Erro", f"Formato {file_format} não implementado ainda.")
            return []
    
    def _process_ofx(self, file_path: str) -> List[Dict]:
        transactions = []
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='latin-1') as file:
                content = file.read()

        transaction_pattern = r'<STMTTRN>(.*?)</STMTTRN>'
        matches = re.findall(transaction_pattern, content, re.DOTALL)

        print(f"Total de transações encontradas no arquivo: {len(matches)}")

        for match in matches:
            trntype_match = re.search(r'<TRNTYPE>(.*?)</TRNTYPE>', match)
            date_match = re.search(r'<DTPOSTED>(\d{8}).*?</DTPOSTED>', match)
            amount_match = re.search(r'<TRNAMT>(-?\d+\.?\d*)', match)
            memo_match = re.search(r'<MEMO>(.*?)</MEMO>', match)

            if all([trntype_match, date_match, amount_match, memo_match]):
                amount = float(amount_match.group(1).strip())
                
                if amount < 0:
                    amount = abs(amount)
                    ofx_date = date_match.group(1).strip()
                    parsed_date = self._parse_ofx_date(ofx_date)
                    memo = memo_match.group(1).strip()
                    fornecedor = self._clean_sicoob_description(memo)
                    
                    transactions.append({
                        'fornecedor': fornecedor,
                        'categoria': 'Cartão de Credito',
                        'valor': amount,
                        'data_registro': parsed_date
                    })
                else:
                    print(f"Ignorando transação com valor positivo: {memo_match.group(1).strip()}")

        return transactions

    def _parse_ofx_date(self, ofx_date: str) -> str:
        date_part = ofx_date[:8]
        try:
            dt = datetime.strptime(date_part, "%Y%m%d")
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            return "01/01/2025"

    def _clean_description(self, description: str) -> str:
        cleaned = re.sub(r'\s+', ' ', description)
        cleaned = cleaned.strip()
        if len(cleaned) > 50:
            cleaned = cleaned[:50] + "..."
        return cleaned

    def _clean_sicoob_description(self, description: str) -> str:
        cleaned = re.sub(r'\s+\d{2}/\d{2}\s+', ' ', description)
        cities = ['RIBEIRAO PRET', 'RIBEIRAO PRE', 'SAO PAULO', 'OSASCO', 'HORTOLANDIA', 
                  'BELO HORIZON', 'SAN FRANCISCO', 'ARIBEIRAO PRE']
        for city in cities:
            cleaned = re.sub(rf'\s*{re.escape(city)}.*$', '', cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r'\s+', ' ', cleaned)
        cleaned = re.sub(r'\s*-?\s*US\$.*$', '', cleaned)
        return cleaned.strip()

    def _extract_text_with_ocr(self, file_path: str) -> str:
        """
        Extrai texto de PDF usando OCR quando necessário.
        """
        text = ""
        try:
            # Tenta primeiro com pdfplumber (para PDFs com texto)
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            
            # Se não conseguiu extrair texto suficiente, tenta OCR
            if len(text.strip()) < 100:
                print("PDF parece ser uma imagem. Aplicando OCR...")
                
                # Usa PyMuPDF para converter páginas em imagens
                pdf_document = fitz.open(file_path)
                
                for page_num in range(len(pdf_document)):
                    page = pdf_document[page_num]
                    pix = page.get_pixmap(matrix=fitz.Matrix(300/72, 300/72))  # 300 DPI
                    img_data = pix.tobytes("png")
                    
                    # Converte para PIL Image
                    from io import BytesIO
                    img = Image.open(BytesIO(img_data))
                    
                    # Aplica OCR
                    try:
                        page_text = pytesseract.image_to_string(img, lang='por')
                        text += page_text + "\n"
                    except Exception as e:
                        print(f"Erro no OCR da página {page_num + 1}: {e}")
                
                pdf_document.close()
                
        except Exception as e:
            print(f"Erro ao processar PDF: {e}")
            messagebox.showerror("Erro", f"Erro ao processar PDF: {e}")
            
        return text

    def _process_pdf(self, file_path: str, bank: str) -> List[Dict]:
        transactions = []
        try:
            # Usa o novo método que suporta OCR
            text = self._extract_text_with_ocr(file_path)
            
            if bank == "Itaú":
                transactions = self._parse_itau_pdf(text)
            elif bank == "Banco do Brasil":
                transactions = self._parse_bb_pdf(text)
            elif bank == "Caixa":
                transactions = self._parse_cef_pdf(text)
            elif bank == "Sicoob":
                transactions = self._parse_sicoob_pdf(text)
            elif bank == "Santander":
                transactions = self._parse_santander_pdf(text)
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar PDF: {e}")
            raise
        return transactions
    
    def _parse_santander_pdf(self, text: str) -> List[Dict]:
        """
        Parser específico para extratos do Santander.
        """
        transactions = []
        lines = text.split('\n')

        # Tenta encontrar o ano da fatura de forma mais robusta
        current_year = datetime.now().year
        year_match = re.search(r'Vencimento\s+\d{2}/\d{2}/(20\d{2})', text, re.IGNORECASE)
        if year_match:
            current_year = int(year_match.group(1))

        # Palavras-chave para ignorar
        ignore_keywords = ['TOTAL', 'SALDO', 'PAGAMENTO', 'FATURA', 'ANTERIOR', 'CRÉDITO', 
                          'DÉBITO AUTOM', 'ENCARGOS', 'ANUIDADE DIFERENCIADA', 'RESUMO',
                          'LIMITE', 'DISPONÍVEL']
        
        # Padrão de regex unificado para capturar transações
        # Captura data (com ou sem ano), descrição e valor, garantindo que "R$"
        # não seja parte da descrição.
        pattern = r'(\d{2}/\d{2}(?:/\d{2,4})?)\s+(.+?)\s+(?:R\$)?\s*([\d.,]+,\d{2})'

        for line in lines:
            line = line.strip()
            if not line:
                continue
                
            # Pula linhas com palavras-chave de resumo ou que sejam de crédito
            if any(keyword in line.upper() for keyword in ignore_keywords):
                continue
            if re.search(r'-\s*[\d.,]+,\d{2}', line): # Ignora valores negativos
                continue
            
            match = re.search(pattern, line)
            
            if match:
                date_str = match.group(1)
                description = match.group(2).strip()
                value_str = match.group(3)

                try:
                    value = float(value_str.replace('.', '').replace(',', '.'))
                    if value > 0:
                        # Limpa a descrição
                        clean_description = self._clean_santander_description(description)

                        # Validação final para evitar descrições vazias ou "R$"
                        if not clean_description or clean_description.upper() == 'R$':
                            continue
                        clean_description = self._clean_santander_description(description)

                        # Formata a data, adicionando o ano se necessário
                        date_parts = date_str.split('/')
                        if len(date_parts) == 2:
                            date_formatted = f"{date_parts[0]}/{date_parts[1]}/{current_year}"
                        else:
                            # Garante que o ano tenha 4 dígitos
                            day, month, year_part = date_parts
                            if len(year_part) == 2:
                                date_formatted = f"{day}/{month}/20{year_part}"
                            else:
                                date_formatted = date_str

                        transactions.append({
                            'fornecedor': clean_description,
                            'categoria': 'Cartão de Credito',
                            'valor': value,
                            'data_registro': date_formatted
                        })
                        
                except ValueError as e:
                    print(f"Erro ao processar a linha (Santander): '{line}'. Erro: {e}")
                    continue
        
        print(f"Total de transações encontradas (Santander): {len(transactions)}")
        return transactions
    
    def _clean_santander_description(self, description: str) -> str:
        """
        Limpa a descrição das transações do Santander.
        """
        # Remove informações de parcelamento (ex: 01/12)
        cleaned = re.sub(r'\s+\d{2}/\d{2}\s*$', '', description).strip()
        # Remove códigos no final (ex: "EMCT06D06")
        cleaned = re.sub(r'\s+[A-Z0-9]{6,8}$', '', cleaned).strip()
        # Remove asteriscos e excesso de espaços
        cleaned = re.sub(r'[*]', ' ', cleaned).strip()
        cleaned = re.sub(r'\s+', ' ', cleaned)

        # Remove "R$" se for a única coisa que sobrou
        if cleaned.upper() == 'R$':
            return ""

        # Capitaliza as palavras para um formato mais limpo
        return ' '.join(word.capitalize() for word in cleaned.split())

    def _parse_itau_pdf(self, text: str) -> List[Dict]:
        transactions = []
        pattern = r"(\d{2}/\d{2})\s+([^\n]+?)\s+R\$?([\d\.]+,\d{2})"
        lines = text.split('\n')
        full_text = ' '.join(lines)
        current_year = datetime.now().year
        year_match = re.search(r'(\d{2}/\d{2}/(\d{4}))', full_text)
        if year_match:
            current_year = int(year_match.group(2))
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if any(word in line for word in ['Total', 'Saldo', 'Pagamento', 'Encargos', 'Tarifas', 'Custo Efetivo']):
                continue
            match = re.search(pattern, line)
            if match:
                date_str = match.group(1).strip()
                description = match.group(2).strip()
                value_str = match.group(3).strip()
                if '-' in value_str:
                    continue
                clean_description = self._clean_itau_description(description)
                try:
                    value = float(value_str.replace('.', '').replace(',', '.'))
                    date_formatted = self._convert_date_itau(date_str, current_year)
                    transactions.append({
                        'fornecedor': clean_description,
                        'categoria': 'Cartão de Credito',
                        'valor': value,
                        'data_registro': date_formatted
                    })
                except ValueError as e:
                    print(f"Erro ao converter valor '{value_str}': {e}")
                    continue
        return transactions

    def _clean_itau_description(self, description: str) -> str:
        cleaned = re.sub(r'\s*\d{2}/\d{2}$', '', description)
        cleaned = re.sub(r'\s*un\d{2}/\d{2}$', '', cleaned)
        cleaned = re.sub(r'\s+', ' ', cleaned)
        return cleaned.strip()

    def _parse_bb_pdf(self, text: str) -> List[Dict]:
        transactions = []
        current_year = datetime.now().year
        year_match = re.search(r'(\d{2}/\d{2}/(\d{4}))', text)
        if year_match:
            current_year = int(year_match.group(2))
        lines = text.split('\n')
        pattern = r"(\d{2}/\d{2})\s+(.*?)\s+([\d\.]+,\d{2})"
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if any(word in line for word in ['LANÇAMENTOS', 'TOTAL', 'FATURA', 'SALDO', 'RESUMO', 'ANTERIOR', 'PARCIAL']):
                continue
            if ' - ' in line or '-R$' in line or 'CRÉDITO' in line.upper() or 'ESTORNO' in line.upper():
                continue
            match = re.search(pattern, line)
            if match:
                date_str = match.group(1).strip()
                description = match.group(2).strip()
                value_str = match.group(3).strip()
                try:
                    value = float(value_str.replace('.', '').replace(',', '.'))
                    date_formatted = self._convert_date_bb(date_str, current_year)
                    clean_description = self._clean_bb_description(description)
                    transactions.append({
                        'fornecedor': clean_description,
                        'categoria': 'Cartão de Credito',
                        'valor': value,
                        'data_registro': date_formatted
                    })
                except ValueError as e:
                    print(f"Erro ao converter valor '{value_str}': {e}")
                    continue
        return transactions

    def _clean_bb_description(self, description: str) -> str:
        cleaned = re.sub(r'\s*PARC\s+\d{2}/\d{2}', '', description, flags=re.IGNORECASE)
        cities = ['RIBEIRAO PRET', 'RIBEIRAO PRE', 'SAO PAULO', 'OSASCO', 'HORTOLANDIA', 
                  'BELO HORIZON', 'SAN FRANCISCO']
        for city in cities:
            pattern = r'\s+' + re.escape(city) + r'\s*$'
            cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r'\s+', ' ', cleaned)
        return cleaned.strip()

    def _convert_date_itau(self, date_str: str, year: int) -> str:
        try:
            parts = date_str.replace(' ', '').split('/')
            day = parts[0].zfill(2)
            month = parts[1].zfill(2)
            return f"{day}/{month}/{year}"
        except:
            return "01/01/2025"
    
    def _convert_date_bb(self, date_str: str, year: int) -> str:
        try:
            parts = date_str.replace(' ', '').split('/')
            day = parts[0].zfill(2)
            month = parts[1].zfill(2)
            return f"{day}/{month}/{year}"
        except:
            return "01/01/2025"

    def _parse_cef_pdf(self, text: str) -> List[Dict]:
        transactions = []
        current_year = datetime.now().year
        year_match = re.search(r'(\d{2}/\d{2}/(\d{4}))', text)
        if year_match:
            current_year = int(year_match.group(2))
        lines = text.split('\n')
        processing_section = False
        current_section = ""
        target_sections = ["ANUIDADE", "COMPRAS", "COMPRAS PARCELADAS"]
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            for section in target_sections:
                if section in line and ("Cartão" in line or section == line):
                    processing_section = True
                    current_section = section
                    break
            if processing_section:
                if any(ignore_section in line for ignore_section in ["OUTROS", "Demonstrativo", "Total final", "Valor total desta fatura", "Total COMPRAS", "Total COMPRAS PARCELADAS"]):
                    processing_section = False
                    current_section = ""
                    continue
                if re.search(r'^[A-Z\s]+\s*\(Cartão\s+\d+\)', line) and not any(section in line for section in target_sections):
                    processing_section = False
                    current_section = ""
                    continue
                if any(header in line for header in ["Data", "Descrição", "Cidade/País", "Valor U$$", "Crédito/Débito", "Total", "Valor Original", "Cotação"]):
                    continue
                pattern = r'(\d{2}/\d{2})\s+(.+?)\s+([A-Z][A-Z\s]*[A-Z])\s+([\d\.]+,\d{2})\s*D\s*$'
                match = re.search(pattern, line)
                if match:
                    date_str = match.group(1).strip()
                    description = match.group(2).strip()
                    city = match.group(3).strip()
                    value_str = match.group(4).strip()
                    try:
                        value = float(value_str.replace('.', '').replace(',', '.'))
                        date_formatted = self._convert_date_cef(date_str, current_year)
                        clean_description = self._clean_cef_description(description)
                        transactions.append({
                            'fornecedor': self._clean_description(clean_description),
                            'categoria': 'Cartão de Credito',
                            'valor': value,
                            'data_registro': date_formatted
                        })
                    except ValueError as e:
                        print(f"Erro ao converter valor '{value_str}': {e}")
                        continue
                elif current_section == "ANUIDADE":
                    anuidade_pattern = r'^([A-Z\s\d/]+?)\s+([\d\.]+,\d{2})\s*D\s*$'
                    match = re.search(anuidade_pattern, line)
                    if match:
                        description = match.group(1).strip()
                        value_str = match.group(2).strip()
                        try:
                            value = float(value_str.replace('.', '').replace(',', '.'))
                            date_formatted = f"01/08/{current_year}"
                            transactions.append({
                                'fornecedor': self._clean_description(description),
                                'categoria': 'Cartão de Credito', 
                                'valor': value,
                                'data_registro': date_formatted
                            })
                        except ValueError as e:
                            print(f"Erro ao processar anuidade '{value_str}': {e}")
                            continue
                else:
                    alt_pattern = r'(\d{2}/\d{2})\s+(.+)\s+([\d\.]+,\d{2})\s*D'
                    match = re.search(alt_pattern, line)
                    if match:
                        date_str = match.group(1).strip()
                        full_description = match.group(2).strip()
                        value_str = match.group(3).strip()
                        desc_parts = full_description.rsplit(' ', 2)
                        if len(desc_parts) >= 2:
                            description = ' '.join(desc_parts[:-1])
                        else:
                            description = full_description
                        try:
                            value = float(value_str.replace('.', '').replace(',', '.'))
                            date_formatted = self._convert_date_cef(date_str, current_year)
                            clean_description = self._clean_cef_description(description)
                            transactions.append({
                                'fornecedor': self._clean_description(clean_description),
                                'categoria': 'Cartão de Credito',
                                'valor': value,
                                'data_registro': date_formatted
                            })
                        except ValueError as e:
                            print(f"Erro no padrão alternativo '{value_str}': {e}")
                            continue
        return transactions

    def _clean_cef_description(self, description: str) -> str:
        description = re.sub(r'\s+\d{2}\s+DE\s+\d{2}', '', description, flags=re.IGNORECASE)
        description = re.sub(r'\s+\d{2}/\d{2}$', '', description)
        description = re.sub(r'\s+\d{1,2}/\s*\d{1,2}', '', description)
        description = re.sub(r'\s+-\s+\d+', '', description)
        description = re.sub(r'\s+\d{6}', '', description)
        description = re.sub(r'\*', ' ', description)
        description = re.sub(r'\s+', ' ', description)
        return description.strip()
    
    def _convert_date_cef(self, date_str: str, year: int) -> str:
        try:
            parts = date_str.replace(' ', '').split('/')
            day = parts[0].zfill(2)
            month = parts[1].zfill(2)
            return f"{day}/{month}/{year}"
        except:
            return "01/01/2025"

    def _parse_sicoob_pdf(self, text: str) -> List[Dict]:
        messagebox.showinfo("Aviso", "Lógica para Sicoob (PDF) ainda não implementada.")
        return []

    def _process_excel(self, file_path: str, bank: str) -> List[Dict]:
        transactions = []
        try:
            df = pd.read_excel(file_path)
            if bank == "Caixa":
                transactions = self._parse_cef_excel(df)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar Excel: {e}")
        return transactions

    def _parse_cef_excel(self, df: pd.DataFrame) -> List[Dict]:
        messagebox.showinfo("Aviso", "Lógica para Caixa (Excel) ainda não implementada.")
        return []
    
    def _insert_into_excel(self, file_path: str, transactions: List[Dict], account: str, due_date: str):
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        start_row = 6
        current_row = start_row

        while worksheet[f'C{current_row}'].value is not None:
            current_row += 1

        for transaction in transactions:
            final_fornecedor = transaction.get('fornecedor_omie') or transaction['fornecedor']
            
            worksheet[f'C{current_row}'] = final_fornecedor
            worksheet[f'D{current_row}'] = transaction['categoria']
            worksheet[f'E{current_row}'] = account
            worksheet[f'F{current_row}'] = transaction['valor']
            worksheet[f'J{current_row}'] = transaction['data_registro']
            worksheet[f'K{current_row}'] = due_date
            current_row += 1
        
        workbook.save(file_path)

class ReconciliationWindow(tk.Toplevel):
    def __init__(self, parent, transactions: List[Dict], omie_suppliers: List[Dict], omie_categories: List[Dict]):
        super().__init__(parent)
        self.title("Conciliação Manual de Fornecedores e Categorias")
        self.state('zoomed')
        self.transient(parent)
        self.grab_set()
        
        self.transactions = transactions
        self.omie_suppliers = omie_suppliers
        self.omie_categories = omie_categories
        
        self.supplier_names = sorted([s.get('nome_fantasia') or s.get('razao_social') for s in omie_suppliers] + ["Cartão de Credito"])
        
        self.category_names = sorted([html.unescape(c.get('descricao')) for c in omie_categories if c.get('descricao') and html.unescape(c.get('descricao')).strip().lower() != 'disponível'])

        self.unreconciled_transactions = [t for t in transactions if not t.get('fornecedor_omie')]
        self.tree_items = {}

        self.create_widgets()
        self.populate_treeview()

    def create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        extrato_frame = ttk.Frame(main_frame, padding="10")
        extrato_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        extrato_title = ttk.Label(extrato_frame, text="Itens do Extrato a Conciliar:", font=("Helvetica", 10, "bold"))
        extrato_title.pack(pady=(0, 10))

        columns = ('data_registro', 'fornecedor', 'fornecedor_omie', 'categoria_omie')
        self.tree = ttk.Treeview(extrato_frame, columns=columns, show='headings')
        self.tree.heading('data_registro', text='Data')
        self.tree.heading('fornecedor', text='Descrição Extrato')
        self.tree.heading('fornecedor_omie', text='Fornecedor Omie')
        self.tree.heading('categoria_omie', text='Categoria Omie')
        
        self.tree.column('data_registro', width=80, anchor=tk.CENTER)
        self.tree.column('fornecedor', width=250)
        self.tree.column('fornecedor_omie', width=250)
        self.tree.column('categoria_omie', width=150)

        scrollbar_y = ttk.Scrollbar(extrato_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(extrato_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(fill=tk.BOTH, expand=True)

        self.tree.bind('<Double-1>', self.on_double_click)
        
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))

        self.supplier_tab = ttk.Frame(self.notebook, padding=10)
        self.category_tab = ttk.Frame(self.notebook, padding=10)
        
        self.notebook.add(self.supplier_tab, text="Fornecedores")
        self.notebook.add(self.category_tab, text="Categorias")

        supplier_title = ttk.Label(self.supplier_tab, text="Fornecedores Omie:", font=("Helvetica", 10, "bold"))
        supplier_title.pack(pady=(0, 5))

        search_frame = ttk.Frame(self.supplier_tab)
        search_label = ttk.Label(search_frame, text="Pesquisar:")
        self.search_entry = ttk.Entry(search_frame)
        search_label.pack(side=tk.LEFT)
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        search_frame.pack(fill=tk.X)

        self.supplier_listbox = tk.Listbox(self.supplier_tab)
        self.supplier_listbox.pack(fill=tk.BOTH, expand=True)

        for name in self.supplier_names:
            self.supplier_listbox.insert(tk.END, name)

        self.search_entry.bind('<KeyRelease>', self.filter_suppliers)
        self.supplier_listbox.bind('<Double-1>', self.on_listbox_double_click)

        category_title = ttk.Label(self.category_tab, text="Categorias Omie:", font=("Helvetica", 10, "bold"))
        category_title.pack(pady=(0, 5))

        search_frame_cat = ttk.Frame(self.category_tab)
        search_label_cat = ttk.Label(search_frame_cat, text="Pesquisar:")
        self.search_entry_cat = ttk.Entry(search_frame_cat)
        search_label_cat.pack(side=tk.LEFT)
        self.search_entry_cat.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        search_frame_cat.pack(fill=tk.X)

        self.category_listbox = tk.Listbox(self.category_tab)
        self.category_listbox.pack(fill=tk.BOTH, expand=True)
        
        for name in self.category_names:
            self.category_listbox.insert(tk.END, name)

        self.search_entry_cat.bind('<KeyRelease>', self.filter_categories)
        self.category_listbox.bind('<Double-1>', self.on_category_listbox_double_click)
        
        button_frame = ttk.Frame(self)
        button_frame.pack(pady=10)
        
        save_button = ttk.Button(button_frame, text="Salvar e Fechar", command=self.save_and_close)
        save_button.pack(side=tk.LEFT, padx=10)

    def populate_treeview(self):
        self.tree.delete(*self.tree.get_children())
        self.tree_items = {}
        for transaction in self.unreconciled_transactions:
            item_id = self.tree.insert('', tk.END, values=(
                transaction['data_registro'],
                transaction['fornecedor'],
                '',
                'Cartão de Credito'
            ))
            self.tree_items[item_id] = transaction

    def on_double_click(self, event):
        selected_item = self.tree.identify_row(event.y)
        column_id = self.tree.identify_column(event.x)
        
        if not selected_item:
            return
        
        if column_id == '#3':
            self.notebook.select(self.supplier_tab)
        elif column_id == '#4':
            self.notebook.select(self.category_tab)
        
        self.tree.selection_set(selected_item)
        
    def filter_suppliers(self, event):
        search_term = self.search_entry.get().lower()
        self.supplier_listbox.delete(0, tk.END)
        for name in self.supplier_names:
            if search_term in name.lower():
                self.supplier_listbox.insert(tk.END, name)

    def filter_categories(self, event):
        search_term = self.search_entry_cat.get().lower()
        self.category_listbox.delete(0, tk.END)
        for name in self.category_names:
            if search_term in name.lower():
                self.category_listbox.insert(tk.END, name)

    def on_listbox_double_click(self, event):
        selected_supplier_index = self.supplier_listbox.curselection()
        if not selected_supplier_index:
            return
            
        selected_supplier = self.supplier_listbox.get(selected_supplier_index[0])
        
        selected_tree_item = self.tree.focus()
        if not selected_tree_item:
            messagebox.showwarning("Aviso", "Por favor, selecione uma linha do extrato para alterar.")
            return

        values = list(self.tree.item(selected_tree_item, 'values'))
        values[2] = selected_supplier
        self.tree.item(selected_tree_item, values=values)
        
        self.tree_items[selected_tree_item]['fornecedor_omie'] = selected_supplier
    
    def on_category_listbox_double_click(self, event):
        selected_category_index = self.category_listbox.curselection()
        if not selected_category_index:
            return
        
        selected_category = self.category_listbox.get(selected_category_index[0])
        
        selected_tree_item = self.tree.focus()
        if not selected_tree_item:
            messagebox.showwarning("Aviso", "Por favor, selecione uma linha do extrato para alterar.")
            return

        values = list(self.tree.item(selected_tree_item, 'values'))
        values[3] = selected_category
        self.tree.item(selected_tree_item, values=values)

        self.tree_items[selected_tree_item]['categoria'] = selected_category
            
    def save_and_close(self):
        self.destroy()

# A classe App foi totalmente refeita para usar uma interface mais bonita e organizada.
# A lógica interna dos métodos foi mantida, mas a forma de construir os elementos visuais
# foi modernizada.
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Automatizador de Extratos e Conciliação Omie")
        self.geometry("600x450")  # Define um tamanho inicial para a janela
        self.processor = ExtractProcessor()
        self.clients = ["Aurora Hotel", "Elias Carnes", "Ipê Amarelo", "Boteco Napoleão"]
        
        # Configurar o estilo para um visual mais moderno
        self.style = ttk.Style(self)
        self.style.theme_use('vista')  # Use 'clam', 'alt', 'default', 'vista'
        self.style.configure('TFrame', background='#f0f0f0')
        self.style.configure('TButton', font=('Helvetica', 10, 'bold'), padding=8)
        self.style.configure('TLabel', font=('Helvetica', 10))
        self.style.configure('TCombobox', font=('Helvetica', 10))
        self.style.configure('TEntry', font=('Helvetica', 10))

        # O layout agora usa um frame central para melhor alinhamento
        main_frame = ttk.Frame(self, padding="20 15 20 15", relief="groove", borderwidth=2)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        title_label = ttk.Label(main_frame, text="Automatizador de Extratos de Cartão", font=("Helvetica", 14, "bold"))
        title_label.pack(pady=(0, 15))

        # Reorganizar campos de entrada em frames de 2 colunas
        self.create_input_field(main_frame, "Selecione o cliente:", self.clients, is_combo=True, var_name='client_combo')
        self.create_input_field(main_frame, "Selecione o banco:", self.processor.supported_banks, is_combo=True, var_name='bank_combo')
        self.create_file_field(main_frame, "Arquivo de Extrato:", var_name='file_entry')
        self.create_input_field(main_frame, "Conta Corrente:", var_name='account_entry')
        self.create_input_field(main_frame, "Data de Vencimento (DD/MM/AAAA):", var_name='due_date_entry')

        # Botões
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)
        
        process_button = ttk.Button(button_frame, text="Processar", command=self.process_data)
        process_button.pack(side=tk.LEFT, padx=10)
        
        exit_button = ttk.Button(button_frame, text="Sair", command=self.destroy)
        exit_button.pack(side=tk.LEFT, padx=10)

        self.status_label = ttk.Label(main_frame, text="", font=("Helvetica", 10), foreground="blue", wraplength=500)
        self.status_label.pack(pady=(10, 0))

        # Binds
        self.client_combo.bind("<<ComboboxSelected>>", self.on_client_selected)

    def create_input_field(self, parent_frame, label_text, values=None, is_combo=False, var_name=''):
        field_frame = ttk.Frame(parent_frame, padding="5 5")
        field_frame.pack(fill=tk.X)
        
        label = ttk.Label(field_frame, text=label_text)
        label.pack(side=tk.LEFT, padx=(0, 10))
        
        if is_combo:
            widget = ttk.Combobox(field_frame, values=values, state="readonly")
            if values:
                widget.set(values[0])
        else:
            widget = ttk.Entry(field_frame)
            
        widget.pack(side=tk.LEFT, fill=tk.X, expand=True)
        setattr(self, var_name, widget)

    def create_file_field(self, parent_frame, label_text, var_name=''):
        field_frame = ttk.Frame(parent_frame, padding="5 5")
        field_frame.pack(fill=tk.X)

        label = ttk.Label(field_frame, text=label_text)
        label.pack(side=tk.LEFT, padx=(0, 10))

        widget = ttk.Entry(field_frame)
        widget.pack(side=tk.LEFT, fill=tk.X, expand=True)
        setattr(self, var_name, widget)

        browse_button = ttk.Button(field_frame, text="Procurar...", command=self.browse_file)
        browse_button.pack(side=tk.LEFT, padx=(5, 0))

    def on_client_selected(self, event):
        # Lógica para quando um cliente é selecionado (se precisar de alguma ação)
        pass

    def browse_file(self):
        filetypes = [("Arquivos de Extrato", "*.ofx *.pdf"), ("Todos os arquivos", "*.*")]
        filename = filedialog.askopenfilename(
            title="Selecione o arquivo de extrato",
            filetypes=filetypes
        )
        if filename:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, filename)
    
    def process_data(self):
        bank = self.bank_combo.get()
        file_path = self.file_entry.get()
        account = self.account_entry.get()
        due_date = self.due_date_entry.get()
        client = self.client_combo.get()

        if not all([bank, file_path, account, due_date, client]):
            messagebox.showwarning("Aviso", "Preencha todos os campos antes de processar.")
            return

        try:
            self.status_label.config(text="Processando e conciliando...", foreground="blue")
            self.update_idletasks()
            
            transactions = self.processor._process_and_reconcile(bank, file_path, client)
            
            if not transactions:
                self.status_label.config(text="Erro ou nenhuma transação para processar.", foreground="red")
                return

            default_category = "Cartão de Credito"
            for t in transactions:
                t['categoria'] = default_category

            unreconciled = [t for t in transactions if not t.get('fornecedor_omie')]
            
            if unreconciled:
                reconciliation_window = ReconciliationWindow(self, transactions, self.processor.omie_suppliers, self.processor.omie_categories)
                self.wait_window(reconciliation_window)
                
            result = self.processor.process_and_save(transactions, account, due_date)
            self.status_label.config(text=result, foreground="green" if "✅" in result else "red")
                
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro: {e}")
            self.status_label.config(text=f"Erro: {e}", foreground="red")

if __name__ == "__main__":
    app = App()
    app.mainloop()