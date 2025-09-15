#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Automatizador de Extratos de Cartão de Crédito
Desenvolvido para inserir dados de extratos bancários em planilha Excel
Preservando toda a formatação original da planilha base.

MAPEAMENTO DE COLUNAS:
- Coluna C: Fornecedor (descrição da compra)
- Coluna D: Categoria (sempre "Cartão de Credito")
- Coluna E: Conta Corrente (informada pelo usuário)
- Coluna F: Valor da Conta (valor da transação)
- Coluna J: Data de Registro (data da compra)
- Coluna K: Data de Vencimento (informada pelo usuário)

Inserção inicia na linha 6.
"""

import os
import re
import sys
from datetime import datetime
from typing import List, Dict, Tuple, Optional
import xml.etree.ElementTree as ET
import shutil

# Importa as bibliotecas do Tkinter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    import PyPDF2
    import pdfplumber
except ImportError as e:
    messagebox.showerror(
        "Erro", 
        f"Erro: Biblioteca necessária não encontrada: {e}\n\n"
        "Instale as dependências com:\n"
        "pip install pandas openpyxl PyPDF2 pdfplumber"
    )
    sys.exit(1)

class ExtractProcessor:
    """Classe principal para processar extratos de cartão de crédito"""

    def __init__(self):
        self.supported_banks = ["Sicoob", "Banco do Brasil", "Caixa", "Itaú"]
        self.file_formats = {
            "Sicoob": "OFX",
            "Banco do Brasil": "PDF",
            "Caixa": "PDF",
            "Itaú": "PDF"
        }

    def process_extract_and_save(self, bank: str, extract_file: str, account: str, due_date: str) -> str:
        """
        Processa o extrato e salva os dados na planilha.
        Retorna o caminho do novo arquivo ou uma mensagem de erro.
        """
        base_file = "C:\\Bitrix24\\Aurora Hotel\\Automação\\Omie_Contas_Pagar_v1_1_5.xlsx"
        if not os.path.exists(base_file):
            return "ERRO: Arquivo base 'Omie_Contas_Pagar_v1_1_5.xlsx' não encontrado!"

        file_format = self.file_formats[bank]
        transactions = self._process_extract(extract_file, file_format, bank)

        if not transactions:
            return "Nenhuma transação encontrada no extrato."

        new_file_path = self._create_new_excel_file(base_file)
        if new_file_path:
            self._insert_into_excel(new_file_path, transactions, account, due_date)
            return f"✅ Processamento concluído! {len(transactions)} transações inseridas.\n\nArquivo atualizado: {new_file_path}"
        else:
            return "Erro ao criar a nova planilha."

    def _create_new_excel_file(self, base_file: str) -> str:
        """Cria uma cópia da planilha base com um nome único e a salva na Área de Trabalho"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_file_name = f"Omie_Contas_Pagar_Atualizada_{timestamp}.xlsx"

            # Obtém o caminho da Área de Trabalho do usuário
            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
            
            # Constroi o caminho completo do novo arquivo na Área de Trabalho
            new_file_path = os.path.join(desktop_path, new_file_name)

            shutil.copy2(base_file, new_file_path)
            return new_file_path
        except Exception as e:
            print(f"Erro ao criar nova planilha: {e}")
            return ""

    def _process_extract(self, file_path: str, file_format: str, bank: str) -> List[Dict]:
        """Processa o arquivo de extrato baseado no formato"""
        # Em Tkinter, a comunicação visual não é via pop-up, mas via atualização de status.
        # Esta função é chamada pela interface, que já lida com o feedback visual.
        
        if file_format == "OFX":
            return self._process_ofx(file_path)
        elif file_format == "PDF":
            return self._process_pdf(file_path, bank)
        elif file_format == "Excel":
            return self._process_excel(file_path, bank)
        else:
            messagebox.showerror("Erro", f"Formato {file_format} não implementado ainda.")
            return []

    def _process_ofx(self, file_path: str) -> List[Dict]:
        """Processa arquivo OFX"""
        transactions = []
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='latin-1') as file:
                content = file.read()

        transaction_pattern = r'<STMTTRN>(.*?)</STMTTRN>'
        matches = re.findall(transaction_pattern, content, re.DOTALL)

        for match in matches:
            trntype_match = re.search(r'<TRNTYPE>(.*?)</TRNTYPE>', match)
            date_match = re.search(r'<DTPOSTED>(\d{8}).*?</DTPOSTED>', match)
            amount_match = re.search(r'<TRNAMT>(-?\d+\.?\d*)', match)
            memo_match = re.search(r'<MEMO>(.*?)</MEMO>', match)

            if all([trntype_match, date_match, amount_match, memo_match]):
                trntype = trntype_match.group(1).strip()
                if trntype in ["DEBIT", "PAYMENT"] or "debit" in memo_match.group(1).lower():
                    amount = float(amount_match.group(1).strip())
                    if amount < 0:
                        amount = abs(amount)
                        ofx_date = date_match.group(1).strip()
                        parsed_date = self._parse_ofx_date(ofx_date)
                        memo = memo_match.group(1).strip()
                        # CORRIGIDO: Chamar a função de limpeza de Sicoob correta
                        fornecedor = self._clean_sicoob_description(memo)
                        transactions.append({
                            'fornecedor': fornecedor,
                            'categoria': 'Cartão de Credito',
                            'valor': amount,
                            'data_registro': parsed_date
                        })
        return transactions

    def _parse_ofx_date(self, ofx_date: str) -> str:
        date_part = ofx_date[:8]
        try:
            dt = datetime.strptime(date_part, "%Y%m%d")
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            return "01/01/2025"

    def _clean_description(self, description: str) -> str:
        cleaned = re.sub(r'\s+', ' ', description)
        cleaned = cleaned.strip()
        if len(cleaned) > 50:
            cleaned = cleaned[:50] + "..."
        return cleaned

    def _clean_sicoob_description(self, description: str) -> str:
        """Limpa descrição do Sicoob removendo parcelas e cidade"""
        # Remove informações de parcelas (formato XX/XX)
        cleaned = re.sub(r'\s+\d{2}/\d{2}\s+', ' ', description)
        
        # Remove cidades conhecidas - adicione mais conforme necessário
        cities = ['RIBEIRAO PRET', 'RIBEIRAO PRE', 'SAO PAULO', 'OSASCO', 'HORTOLANDIA', 
                  'BELO HORIZON', 'SAN FRANCISCO', 'ARIBEIRAO PRE']
        for city in cities:
            cleaned = re.sub(rf'\s*{re.escape(city)}.*$', '', cleaned, flags=re.IGNORECASE)
        
        # Remove espaços múltiplos
        cleaned = re.sub(r'\s+', ' ', cleaned)
        
        # Remove informações de câmbio (U$ XX,XX V.DOL X,XXXX)
        cleaned = re.sub(r'\s*-?\s*US\$.*$', '', cleaned)
        
        return cleaned.strip()

    def _process_pdf(self, file_path: str, bank: str) -> List[Dict]:
        transactions = []
        try:
            with pdfplumber.open(file_path) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() + "\n"

            if bank == "Itaú":
                transactions = self._parse_itau_pdf(text)
            elif bank == "Banco do Brasil":
                transactions = self._parse_bb_pdf(text)
            elif bank == "Caixa":
                transactions = self._parse_cef_pdf(text)
            elif bank == "Sicoob":
                transactions = self._parse_sicoob_pdf(text)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar PDF: {e}")
            raise
        return transactions

    def _parse_itau_pdf(self, text: str) -> List[Dict]:
        transactions = []
        pattern = r"(\d{2}/\d{2})\s+([^\n]+?)\s+R\$?([\d\.]+,\d{2})"
        lines = text.split('\n')
        full_text = ' '.join(lines)
        current_year = datetime.now().year
        year_match = re.search(r'(\d{2}/\d{2}/(\d{4}))', full_text)
        if year_match:
            current_year = int(year_match.group(2))
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            if any(word in line for word in ['Total', 'Saldo', 'Pagamento', 'Encargos', 'Tarifas', 'Custo Efetivo']):
                continue

            match = re.search(pattern, line)
            if match:
                date_str = match.group(1).strip()
                description = match.group(2).strip()
                value_str = match.group(3).strip()
                
                if '-' in value_str:
                    continue
                
                clean_description = self._clean_itau_description(description)
                try:
                    value = float(value_str.replace('.', '').replace(',', '.'))
                    date_formatted = self._convert_date_itau(date_str, current_year)
                    transactions.append({
                        'fornecedor': clean_description,
                        'categoria': 'Cartão de Credito',
                        'valor': value,
                        'data_registro': date_formatted
                    })
                except ValueError as e:
                    print(f"Erro ao converter valor '{value_str}': {e}")
                    continue
        return transactions

    def _clean_itau_description(self, description: str) -> str:
        """Limpa descrição do Itaú removendo parcelas"""
        # Remove informações de parcelas no formato XX/XX no final
        cleaned = re.sub(r'\s*\d{2}/\d{2}$', '', description)
        
        # Remove informações de parcelas no formato unXX/XX
        cleaned = re.sub(r'\s*un\d{2}/\d{2}$', '', cleaned)
        
        # Remove espaços múltiplos
        cleaned = re.sub(r'\s+', ' ', cleaned)
        
        return cleaned.strip()

    def _parse_bb_pdf(self, text: str) -> List[Dict]:
        transactions = []
        current_year = datetime.now().year
        year_match = re.search(r'(\d{2}/\d{2}/(\d{4}))', text)
        if year_match:
            current_year = int(year_match.group(2))

        lines = text.split('\n')
        pattern = r"(\d{2}/\d{2})\s+(.*?)\s+([\d\.]+,\d{2})"
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if any(word in line for word in ['LANÇAMENTOS', 'TOTAL', 'FATURA', 'SALDO', 'RESUMO', 'ANTERIOR', 'PARCIAL']):
                continue
            if ' - ' in line or '-R$' in line or 'CRÉDITO' in line.upper() or 'ESTORNO' in line.upper():
                continue
            match = re.search(pattern, line)
            if match:
                date_str = match.group(1).strip()
                description = match.group(2).strip()
                value_str = match.group(3).strip()

                try:
                    value = float(value_str.replace('.', '').replace(',', '.'))
                    date_formatted = self._convert_date_bb(date_str, current_year)
                    clean_description = self._clean_bb_description(description)
                    transactions.append({
                        'fornecedor': clean_description,
                        'categoria': 'Cartão de Credito',
                        'valor': value,
                        'data_registro': date_formatted
                    })
                except ValueError as e:
                    print(f"Erro ao converter valor '{value_str}': {e}")
                    continue
        return transactions

    def _clean_bb_description(self, description: str) -> str:
        """Limpa descrição do Banco do Brasil removendo parcelas e cidade"""
        # Remove informações de parcelas (PARC XX/XX)
        cleaned = re.sub(r'\s*PARC\s+\d{2}/\d{2}', '', description, flags=re.IGNORECASE)
        
        # Lista de cidades conhecidas para remover apenas do final
        cities = ['RIBEIRAO PRET', 'RIBEIRAO PRE', 'SAO PAULO', 'OSASCO', 'HORTOLANDIA', 
                  'BELO HORIZON', 'SAN FRANCISCO']
        
        # Remove cidade apenas se ela aparecer no final da string
        for city in cities:
            # Padrão que garante que a cidade está no final (pode ter espaços extras depois)
            pattern = r'\s+' + re.escape(city) + r'\s*$'
            cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)
        
        # Remove espaços múltiplos
        cleaned = re.sub(r'\s+', ' ', cleaned)
        
        return cleaned.strip()

    def _convert_date_itau(self, date_str: str, year: int) -> str:
        try:
            parts = date_str.replace(' ', '').split('/')
            day = parts[0].zfill(2)
            month = parts[1].zfill(2)
            return f"{day}/{month}/{year}"
        except:
            return "01/01/2025"
    
    def _convert_date_bb(self, date_str: str, year: int) -> str:
        try:
            parts = date_str.replace(' ', '').split('/')
            day = parts[0].zfill(2)
            month = parts[1].zfill(2)
            return f"{day}/{month}/{year}"
        except:
            return "01/01/2025"

    def _parse_cef_pdf(self, text: str) -> List[Dict]:
        transactions = []
        current_year = datetime.now().year
        year_match = re.search(r'(\d{2}/\d{2}/(\d{4}))', text)
        if year_match:
            current_year = int(year_match.group(2))
        lines = text.split('\n')
        processing_section = False
        current_section = ""
        target_sections = ["ANUIDADE", "COMPRAS", "COMPRAS PARCELADAS"]
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            for section in target_sections:
                if section in line and ("Cartão" in line or section == line):
                    processing_section = True
                    current_section = section
                    break
            if processing_section:
                if any(ignore_section in line for ignore_section in ["OUTROS", "Demonstrativo", "Total final", "Valor total desta fatura", "Total COMPRAS", "Total COMPRAS PARCELADAS"]):
                    processing_section = False
                    current_section = ""
                    continue
                if re.search(r'^[A-Z\s]+\s*\(Cartão\s+\d+\)', line) and not any(section in line for section in target_sections):
                    processing_section = False
                    current_section = ""
                    continue
                if any(header in line for header in ["Data", "Descrição", "Cidade/País", "Valor U$$", "Crédito/Débito", "Total", "Valor Original", "Cotação"]):
                    continue
                pattern = r'(\d{2}/\d{2})\s+(.+?)\s+([A-Z][A-Z\s]*[A-Z])\s+([\d\.]+,\d{2})\s*D\s*$'
                match = re.search(pattern, line)
                if match:
                    date_str = match.group(1).strip()
                    description = match.group(2).strip()
                    city = match.group(3).strip()
                    value_str = match.group(4).strip()
                    try:
                        value = float(value_str.replace('.', '').replace(',', '.'))
                        date_formatted = self._convert_date_cef(date_str, current_year)
                        clean_description = self._clean_cef_description(description)
                        transactions.append({
                            'fornecedor': self._clean_description(clean_description),
                            'categoria': 'Cartão de Credito',
                            'valor': value,
                            'data_registro': date_formatted
                        })
                    except ValueError as e:
                        print(f"Erro ao converter valor '{value_str}': {e}")
                        continue
                elif current_section == "ANUIDADE":
                    anuidade_pattern = r'^([A-Z\s\d/]+?)\s+([\d\.]+,\d{2})\s*D\s*$'
                    match = re.search(anuidade_pattern, line)
                    if match:
                        description = match.group(1).strip()
                        value_str = match.group(2).strip()
                        try:
                            value = float(value_str.replace('.', '').replace(',', '.'))
                            date_formatted = f"01/08/{current_year}"
                            transactions.append({
                                'fornecedor': self._clean_description(description),
                                'categoria': 'Cartão de Credito', 
                                'valor': value,
                                'data_registro': date_formatted
                            })
                        except ValueError as e:
                            print(f"Erro ao processar anuidade '{value_str}': {e}")
                            continue
                else:
                    alt_pattern = r'(\d{2}/\d{2})\s+(.+)\s+([\d\.]+,\d{2})\s*D'
                    match = re.search(alt_pattern, line)
                    if match:
                        date_str = match.group(1).strip()
                        full_description = match.group(2).strip()
                        value_str = match.group(3).strip()
                        desc_parts = full_description.rsplit(' ', 2)
                        if len(desc_parts) >= 2:
                            description = ' '.join(desc_parts[:-1])
                        else:
                            description = full_description
                        try:
                            value = float(value_str.replace('.', '').replace(',', '.'))
                            date_formatted = self._convert_date_cef(date_str, current_year)
                            clean_description = self._clean_cef_description(description)
                            transactions.append({
                                'fornecedor': self._clean_description(clean_description),
                                'categoria': 'Cartão de Credito',
                                'valor': value,
                                'data_registro': date_formatted
                            })
                        except ValueError as e:
                            print(f"Erro no padrão alternativo '{value_str}': {e}")
                            continue
        return transactions

    def _clean_cef_description(self, description: str) -> str:
        description = re.sub(r'\s+\d{2}\s+DE\s+\d{2}', '', description, flags=re.IGNORECASE)
        description = re.sub(r'\s+\d{2}/\d{2}$', '', description)
        description = re.sub(r'\s+\d{1,2}/\s*\d{1,2}', '', description)
        description = re.sub(r'\s+-\s+\d+', '', description)
        description = re.sub(r'\s+\d{6}', '', description)
        description = re.sub(r'\*', ' ', description)
        description = re.sub(r'\s+', ' ', description)
        return description.strip()
    
    def _convert_date_cef(self, date_str: str, year: int) -> str:
        try:
            parts = date_str.replace(' ', '').split('/')
            day = parts[0].zfill(2)
            month = parts[1].zfill(2)
            return f"{day}/{month}/{year}"
        except:
            return "01/01/2025"

    def _parse_sicoob_pdf(self, text: str) -> List[Dict]:
        messagebox.showinfo("Aviso", "Lógica para Sicoob (PDF) ainda não implementada.")
        return []

    def _process_excel(self, file_path: str, bank: str) -> List[Dict]:
        transactions = []
        try:
            df = pd.read_excel(file_path)
            if bank == "Caixa":
                transactions = self._parse_cef_excel(df)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar Excel: {e}")
        return transactions

    def _parse_cef_excel(self, df: pd.DataFrame) -> List[Dict]:
        messagebox.showinfo("Aviso", "Lógica para Caixa (Excel) ainda não implementada.")
        return []

    def _insert_into_excel(self, file_path: str, transactions: List[Dict], account: str, due_date: str):
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        start_row = 6
        current_row = start_row

        while worksheet[f'C{current_row}'].value is not None:
            current_row += 1

        for transaction in transactions:
            worksheet[f'C{current_row}'] = transaction['fornecedor']
            worksheet[f'D{current_row}'] = transaction['categoria']
            worksheet[f'E{current_row}'] = account
            worksheet[f'F{current_row}'] = transaction['valor']
            worksheet[f'J{current_row}'] = transaction['data_registro']
            worksheet[f'K{current_row}'] = due_date
            current_row += 1
        workbook.save(file_path)

# --- Fim da classe de processamento, início da classe da interface ---

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Automatizador de Extratos")
        self.geometry("500x400")
        self.processor = ExtractProcessor()
        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Título
        title_label = ttk.Label(main_frame, text="Automatizador de Extratos de Cartão", font=("Helvetica", 16, "bold"))
        title_label.pack(pady=10)

        # Seleção de Banco
        bank_frame = ttk.Frame(main_frame)
        bank_label = ttk.Label(bank_frame, text="Selecione o banco:")
        self.bank_combo = ttk.Combobox(bank_frame, values=self.processor.supported_banks, state="readonly")
        self.bank_combo.current(0)
        bank_label.pack(side=tk.LEFT, padx=5, pady=5)
        self.bank_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=5)
        bank_frame.pack(fill=tk.X)

        # Caminho do Arquivo
        file_frame = ttk.Frame(main_frame)
        file_label = ttk.Label(file_frame, text="Arquivo de Extrato:")
        self.file_entry = ttk.Entry(file_frame)
        browse_button = ttk.Button(file_frame, text="Procurar...", command=self.browse_file)
        file_label.pack(side=tk.LEFT, padx=5, pady=5)
        self.file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=5)
        browse_button.pack(side=tk.LEFT, padx=5, pady=5)
        file_frame.pack(fill=tk.X)

        # Conta Corrente
        account_frame = ttk.Frame(main_frame)
        account_label = ttk.Label(account_frame, text="Conta Corrente:")
        self.account_entry = ttk.Entry(account_frame)
        account_label.pack(side=tk.LEFT, padx=5, pady=5)
        self.account_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=5)
        account_frame.pack(fill=tk.X)

        # Data de Vencimento
        due_date_frame = ttk.Frame(main_frame)
        due_date_label = ttk.Label(due_date_frame, text="Data de Vencimento (DD/MM/AAAA):")
        self.due_date_entry = ttk.Entry(due_date_frame)
        due_date_label.pack(side=tk.LEFT, padx=5, pady=5)
        self.due_date_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=5)
        due_date_frame.pack(fill=tk.X)

        # Botões
        button_frame = ttk.Frame(main_frame)
        process_button = ttk.Button(button_frame, text="Processar", command=self.process_data)
        exit_button = ttk.Button(button_frame, text="Sair", command=self.destroy)
        process_button.pack(side=tk.LEFT, padx=10, pady=20)
        exit_button.pack(side=tk.LEFT, padx=10, pady=20)
        button_frame.pack()

        # Saída de Status
        self.status_label = ttk.Label(main_frame, text="", font=("Helvetica", 10), wraplength=450)
        self.status_label.pack(pady=10)

    def browse_file(self):
        filetypes = [("Arquivos de Extrato", "*.ofx *.pdf"), ("Todos os arquivos", "*.*")]
        filename = filedialog.askopenfilename(
            title="Selecione o arquivo de extrato",
            filetypes=filetypes
        )
        if filename:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, filename)

    def process_data(self):
        bank = self.bank_combo.get()
        file_path = self.file_entry.get()
        account = self.account_entry.get()
        due_date = self.due_date_entry.get()

        if not all([bank, file_path, account, due_date]):
            messagebox.showwarning("Aviso", "Preencha todos os campos antes de processar.")
            return

        try:
            self.status_label.config(text="Processando...")
            self.update_idletasks()
            
            result = self.processor.process_extract_and_save(bank, file_path, account, due_date)
            
            self.status_label.config(text=result)
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro: {e}")
            self.status_label.config(text=f"Erro: {e}")

if __name__ == "__main__":
    app = App()
    app.mainloop()